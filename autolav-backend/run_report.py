#!/usr/bin/env python3
"""
Script para executar relatório via linha de comando

Uso:
    python run_report.py --units 101,102,103 --start 2025-01-01 --end 2025-01-07 --output report.csv
"""
import asyncio
import argparse
import csv
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook

from app.scraper import LavanderiaPortalScraper
from app.parser import validate_date_range
from app.config import settings
from app.logger import logger


async def run_report(
    units: list,
    start_date: str,
    end_date: str,
    output_format: str = 'csv',
    output_path: str = None
):
    """
    Executa relatório e salva em arquivo
    
    Args:
        units: Lista de IDs de unidades
        start_date: Data inicial
        end_date: Data final
        output_format: Formato de saída ('csv' ou 'excel')
        output_path: Caminho do arquivo de saída
    """
    logger.info(f"Iniciando relatório para {len(units)} unidades")
    logger.info(f"Período: {start_date} a {end_date}")
    
    # Valida datas
    start_date, end_date = validate_date_range(start_date, end_date)
    
    # Cria scraper
    async with LavanderiaPortalScraper() as scraper:
        results = []
        
        for unit_id in units:
            logger.info(f"Processando unidade {unit_id}...")
            
            result = await scraper.scrape_unit(
                unit_id=unit_id,
                start_date=start_date,
                end_date=end_date
            )
            
            results.append(result)
            
            # Delay entre unidades
            await asyncio.sleep(settings.nav_delay / 1000)
    
    # Gera arquivo de saída
    if not output_path:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = settings.reports_dir / f"relatorio_{timestamp}.{output_format}"
    else:
        output_path = Path(output_path)
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    if output_format == 'csv':
        save_csv(results, output_path)
    elif output_format == 'excel':
        save_excel(results, output_path)
    else:
        raise ValueError(f"Formato não suportado: {output_format}")
    
    logger.info(f"Relatório salvo em: {output_path}")
    
    # Mostra resumo
    print("\n" + "="*60)
    print("RESUMO DO RELATÓRIO")
    print("="*60)
    
    for result in results:
        status = "✓" if not result.get('error') else "✗"
        print(f"{status} Unidade {result['unit_id']}: {result['total']} kg ({len(result['rows'])} dias)")
        if result.get('error'):
            print(f"  Erro: {result['error']}")
    
    total_geral = sum(r['total'] for r in results if not r.get('error'))
    print("="*60)
    print(f"TOTAL GERAL: {total_geral} kg")
    print("="*60 + "\n")


def save_csv(results: list, output_path: Path):
    """Salva resultados em CSV"""
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        
        # Cabeçalho
        writer.writerow(['Unidade', 'Data', 'Kg', 'Total Unidade', 'Erro'])
        
        # Dados
        for result in results:
            unit_id = result['unit_id']
            total = result['total']
            error = result.get('error', '')
            
            if result['rows']:
                for row in result['rows']:
                    writer.writerow([
                        unit_id,
                        row['date'],
                        row['kg'],
                        total,
                        error
                    ])
            else:
                writer.writerow([unit_id, '', '', total, error])


def save_excel(results: list, output_path: Path):
    """Salva resultados em Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    
    # Cabeçalho
    headers = ['Unidade', 'Data', 'Kg', 'Total Unidade', 'Erro']
    ws.append(headers)
    
    # Estilo do cabeçalho
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Dados
    for result in results:
        unit_id = result['unit_id']
        total = result['total']
        error = result.get('error', '')
        
        if result['rows']:
            for row in result['rows']:
                ws.append([
                    unit_id,
                    row['date'],
                    row['kg'],
                    total,
                    error
                ])
        else:
            ws.append([unit_id, '', '', total, error])
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Gera relatório de kg por unidade")
    parser.add_argument('--units', required=True, help='IDs das unidades separados por vírgula')
    parser.add_argument('--start', required=True, help='Data inicial (YYYY-MM-DD)')
    parser.add_argument('--end', required=True, help='Data final (YYYY-MM-DD)')
    parser.add_argument('--format', choices=['csv', 'excel'], default='csv', help='Formato de saída')
    parser.add_argument('--output', help='Caminho do arquivo de saída')
    
    args = parser.parse_args()
    
    units = [u.strip() for u in args.units.split(',')]
    
    asyncio.run(run_report(
        units=units,
        start_date=args.start,
        end_date=args.end,
        output_format=args.format,
        output_path=args.output
    ))


if __name__ == "__main__":
    main()
